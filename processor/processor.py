import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
import shutil  # Added for moving files

class TicketProcessor:
    def __init__(self):
        pass

    def filter_last_8_days(self, df, date_column='Ticket Creation Date', current_date=None):
        """
        Filters the DataFrame to include rows where the date in the specified column is within the last 8 days.
        """
        try:
            if current_date is None:
                current_date = pd.to_datetime('today')

            df[date_column] = pd.to_datetime(df[date_column], errors='coerce')
            filtered_df = df[df[date_column] >= (current_date - pd.Timedelta(days=8))]

            return filtered_df
        except Exception as e:
            print(f"Error: {e}")
            return pd.DataFrame()

    def split_dataframe_by_bot_name(self, df, bot_name_list, bot_name_column='Bot Name'):
        """
        Splits the DataFrame into a dictionary of DataFrames based on unique values in the specified bot name column.
        """
        try:
            if bot_name_column not in df.columns:
                raise ValueError(f"Column '{bot_name_column}' does not exist in the DataFrame.")

            splitted_dfs = {bot: df[df[bot_name_column] == bot] for bot in bot_name_list}

            return splitted_dfs
        except Exception as e:
            print(f"Error: {e}")
            return {}, []

    def total_ticket_received_count(self, splitted_df_dict, splitted_df_key_list, comparison_values, column_name='Mode'):
        """
        Count occurrences of specified values in a specified column for a dictionary of DataFrames.
        """
        all_counts = {}

        for key in splitted_df_key_list:
            if key not in splitted_df_dict:
                print(f"Warning: Key '{key}' does not exist in the dictionary. Skipping...")
                continue

            df = splitted_df_dict[key]

            if column_name not in df.columns:
                print(f"Warning: Column '{column_name}' does not exist in DataFrame '{key}'. Skipping...")
                continue

            for value in comparison_values:
                count = (df[column_name] == value).sum()
                all_counts[f"{key}_{value}"] = count if count > 0 else 0

        return all_counts

    def count_tickets_for_unique_modes(self, splitted_df_dict, splitted_df_key_list, mode_list, status_column, status_value):
        """
        Count occurrences of each unique value in a specified column for rows where the status equals a given value.
        """
        all_counts = {}

        for key in splitted_df_key_list:
            if key not in splitted_df_dict:
                print(f"Warning: Key '{key}' does not exist in the dictionary. Skipping...")
                continue

            df = splitted_df_dict[key]
            filtered_df = df[df[status_column] == status_value]

            for value in mode_list:
                count = (filtered_df['Mode'] == value).sum()
                all_counts[f'total_{status_value}_ticket_{key}_{value}'] = count if count > 0 else 0

        return all_counts
    
    def total_ticket_received_count_modewise(self, splitted_df_dict, splitted_df_key_list, comparison_values, column_name):
        """
        Count occurrences of specified values in a specified column for a dictionary of DataFrames.
        The result is stored in separate dictionaries for each unique mode.

        Parameters:
        splitted_df_dict (dict): A dictionary of DataFrames to process, with keys as identifiers.
        splitted_df_key_list (list): A list of keys to specify which DataFrames to process.
        comparison_values (list): A list of values (modes) to count in the specified column.
        column_name (str): The name of the column to check (e.g., 'Mode').

        Returns:
        dict: A dictionary where each mode has its own sub-dictionary containing counts for each DataFrame.
        """
        # Initialize an empty dictionary to store results for each mode
        mode_counts = {value: {} for value in comparison_values}

        # Loop through each key in the provided key list
        for key in splitted_df_key_list:
            if key not in splitted_df_dict.keys():
                print(f"Warning: Key '{key}' does not exist in the dictionary. Skipping...")
                continue

            df = splitted_df_dict[key]  # Get the DataFrame associated with the key

            # Check if the column exists in the current DataFrame
            if column_name not in df.columns:
                print(f"Warning: Column '{column_name}' does not exist in DataFrame '{key}'. Skipping...")
                continue

            # Loop through each value (mode) in the comparison values list
            for value in comparison_values:
                # Count occurrences of the value in the specified column for the current DataFrame
                count = (df[column_name] == value).sum()

                # Store the count in the respective mode's dictionary
                mode_counts[value][key] = count

                # Set count to 0 if the value does not exist in the current DataFrame
                if count == 0:
                    mode_counts[value][key] = 0

        return mode_counts

    def problem_reported_count_botwise(self, splitted_df_dict, splitted_df_key_list, comparison_values, column_name):
        """
        Count occurrences of specified values in a specified column for a dictionary of DataFrames.
        The result is stored in separate dictionaries for each unique mode.

        Parameters:
        splitted_df_dict (dict): A dictionary of DataFrames to process, with keys as identifiers.
        splitted_df_key_list (list): A list of keys to specify which DataFrames to process.
        comparison_values (list): A list of values (modes) to count in the specified column.
        column_name (str): The name of the column to check (e.g., 'Mode').

        Returns:
        dict: A dictionary where each mode has its own sub-dictionary containing counts for each DataFrame.
        """
        # Initialize an empty dictionary to store results for each mode
        mode_counts = {value: {} for value in comparison_values}

        # Loop through each key in the provided key list
        for bot_key in splitted_df_key_list:
            if bot_key not in splitted_df_dict.keys():
                print(f"Warning: Key '{bot_key}' does not exist in the dictionary. Skipping...")
                continue

            df = splitted_df_dict[bot_key]  # Get the DataFrame associated with the key
            # print(df.head(10))

            # Check if the column exists in the current DataFrame
            if column_name not in df.columns:
                print(f"Warning: Column '{column_name}' does not exist in DataFrame '{key}'. Skipping...")
                continue

            # Clean the column (lowercase and strip)
            df.loc[:, column_name] = df[column_name].str.lower().str.strip()
            # df[column_name] = df[column_name].str.lower().str.strip()

            # Loop through each value (mode) in the comparison values list
            for value in comparison_values:
                # Clean the comparison value (lowercase and strip)
                cleaned_value = value.lower().strip()
                # print(cleaned_value)

                # Count occurrences of the cleaned value in the specified column for the current DataFrame
                count = (df[column_name] == cleaned_value).sum()
                # print(count)

                # Store the count in the respective mode's dictionary
                mode_counts[value][bot_key] = count
                # print(mode_counts)

                # print(f"Problem: {value}, Bot Name: {bot_key}, Count: {count}")  # Debugging print

        return mode_counts

    def create_excel_with_pivot_and_dicts(self, mode_wise_data_dict, dict_mapping, problem_reported_botwise_df , file_path):
        """
        Creates an Excel file with a mode-wise pivot at the top followed by multiple dictionaries.

        Parameters:
        mode_wise_data_dict (dict): A dictionary containing mode-wise ticket counts.
        dict_mapping (dict): A dictionary where keys are section names and values are dictionaries of bot ticket counts.
        file_path (str): The file path where the Excel file will be saved.
        
        Returns:
        None
        """
        # Create a new workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active

        # Define bold fonts for different headers
        bold_font = Font(bold=True, size=14)
        medium_font = Font(bold=True, size=12)

        # Part 1: Write the mode-wise pivot data (at the top)
        # Write the header for the mode-wise pivot section
        ws.append(['Total Tickets Received (Current Week Mode wise)'])
        ws.cell(row=ws.max_row, column=1).font = bold_font

        # Define the header for the pivot (bold)
        header = ['Bot Name'] + list(mode_wise_data_dict.keys())  # ['Bot Name', 'Email', 'Phone', 'Web', 'Chat']
        ws.append(header)

        # Apply the medium font to the pivot header row
        for col_num in range(1, len(header) + 1):
            ws.cell(row=ws.max_row, column=col_num).font = medium_font

        # Write the rows with bot names and their corresponding ticket counts for each mode
        bot_names = list(mode_wise_data_dict[next(iter(mode_wise_data_dict))].keys())  # Get bot names from any mode (assuming all modes have same bot names)
        for bot_name in bot_names:
            row = [bot_name] + [mode_wise_data_dict[mode].get(bot_name, 0) for mode in mode_wise_data_dict]
            ws.append(row)

        # Add a blank row after the pivot table
        ws.append([])

        # Part 2: Write multiple dictionaries with unique headers (after pivot table)
        for name, d in dict_mapping.items():
            # Write the variable name as the header for each dictionary (bold)
            ws.append([name])  # Add the dictionary name as a header
            ws.cell(row=ws.max_row, column=1).font = bold_font

            # Write the general header (bold)
            ws.append(['Bot Name', 'Ticket Count'])
            ws.cell(row=ws.max_row, column=1).font = medium_font
            ws.cell(row=ws.max_row, column=2).font = medium_font

            # Write the dictionary data (each key-value pair)
            for bot_name, ticket_count in d.items():
                ws.append([bot_name, ticket_count])
            
            # Add a blank row after each dictionary
            ws.append([])

        # Part 3: Write the output DataFrame at the end
        if not problem_reported_botwise_df.empty:
            # Add a new worksheet for the DataFrame
            ws_output = wb.create_sheet(title='Problem Reported Details')

            # Write the DataFrame header
            # Write the DataFrame header with a margin (one cell padding)
            header = ['Problem Reported'] + list(problem_reported_botwise_df.columns)  # Add an empty string for padding
            ws_output.append(header)
            for col_num in range(1, len(header) + 1):
                ws_output.cell(row=1, column=col_num).font = medium_font

            # Write the DataFrame data
            for row in problem_reported_botwise_df.itertuples(index=True):
                ws_output.append(row)
        
        # Save the workbook to an Excel file
        wb.save(file_path)
        print(f"Excel file saved to: {file_path}")

    def count_bot_wise_status_tickets(self, splitted_df_dict, splitted_df_key_list, mode_list, status_column, status_value):
        """
        Count the total number of tickets for each bot where the status equals a given value based on specified modes.
        """
        total_ticket_counts = {}

        for key in splitted_df_key_list:
            if key not in splitted_df_dict:
                print(f"Warning: Key '{key}' does not exist in the dictionary. Skipping...")
                continue

            df = splitted_df_dict[key]
            filtered_df = df[df[status_column] == status_value]

            total_count = sum((filtered_df['Mode'] == mode).sum() for mode in mode_list)
            total_ticket_counts[key] = total_count

        return total_ticket_counts

    def filter_before_last_8_days(self, df, date_column='Ticket Creation Date', current_date=None):
        """
        Filters the DataFrame to include only rows where the date in the specified column is before the last 8 days.
        """
        try:
            if current_date is None:
                current_date = pd.to_datetime('today')

            df[date_column] = pd.to_datetime(df[date_column], errors='coerce')
            cutoff_date = current_date - pd.Timedelta(days=8)
            before_last_8_days_df = df[df[date_column] < cutoff_date]

            return before_last_8_days_df
        except Exception as e:
            print(f"Error: {e}")
            return pd.DataFrame()
